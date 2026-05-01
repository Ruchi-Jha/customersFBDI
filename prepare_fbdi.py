# =============================================================================
# Oracle Fusion AR — Customer FBDI Preparation Script  v2
#
# Usage:
#   python prepare_fbdi_v2.py --batch-map batch_identifiers.csv [--group <GROUP_ID>]
#   python prepare_fbdi_v2.py --batch-map batch_identifiers.csv --list-groups
#   python prepare_fbdi_v2.py --batch-map batch_identifiers.csv --parallel
#
# Example:
#   python prepare_fbdi_v2.py --batch-map batch_identifiers.csv --group 0001_BPS
#   python prepare_fbdi_v2.py --batch-map batch_identifiers.csv --list-groups
#
# Folder structure expected:
#   customer_fbdi/
#   ├── prepare_fbdi_v2.py          <- this script
#   ├── source_data/
#   │   ├── party.csv
#   │   ├── locations.csv
#   │   ├── party_sites.csv
#   │   ├── accounts.csv
#   │   ├── account_sites.csv
#   │   └── profile_class.csv
#   ├── templates/
#   │   └── CustomerImportTemplate.xlsm   <- Standard Oracle FBDI template
#   └── output/
#       └── CustomerImportTemplate_BATCH001.xlsm  <- produced by this script
#
# What this script does:
#   1. Reads source data from 6 CSV files
#   2. Validates OriginalSystemReference columns (not blank + parent-child match)
#   3. Maps and transforms data into the correct FBDI sheet columns
#   4. Writes results into a COPY of the Oracle FBDI .xlsm template
#   5. Stamps the Batch Identifier into every sheet
#   6. Saves the output as CustomerImportTemplate_<BATCH_ID>.xlsm
#
# Dependencies:
#   pip install pandas openpyxl
# =============================================================================

import argparse
import concurrent.futures
import json
import os
import re
import shutil
import sys
from datetime import datetime

import openpyxl
import pandas as pd


# =============================================================================
# CONFIGURATION — adjust these if your folder names or file names differ
# =============================================================================

SOURCE_DIR   = "source_data"
TEMPLATE_DIR = "templates"
OUTPUT_DIR   = "output"
CONFIG_FILE  = "source_config.json"
BATCH_MAP_FILE = "batch_identifiers.csv"

DEFAULT_SOURCE_FILE_PREFIXES = {
    "party"        : "CUST_PARTY",
    "locations"    : "CUST_LOC",
    "party_sites"  : "CUST_PARTY_SITE",
    "accounts"     : "CUST_ACCT",
    "account_sites": "CUST_ACCT_SITE",
    "profile_class": "CUST_PROF",
}

SOURCE_KEYS = list(DEFAULT_SOURCE_FILE_PREFIXES.keys())

# Name of the Oracle standard FBDI template file inside templates/
TEMPLATE_FILENAME = "CustomerImportTemplate.xlsm"


def load_source_file_config():
    """Load source configuration for dynamic file discovery from JSON."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, CONFIG_FILE)

    if not os.path.exists(config_path):
        abort(
            f"Source config file not found: {os.path.abspath(config_path)}\n"
            f"        Please create '{CONFIG_FILE}' in the script folder."
        )

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception as exc:
        abort(f"Failed to read '{CONFIG_FILE}': {exc}")

    if not isinstance(config, dict):
        abort(f"'{CONFIG_FILE}' must contain a JSON object at the top level.")

    source_dir = config.get("source_dir", SOURCE_DIR)
    if not isinstance(source_dir, str) or not source_dir.strip():
        abort(f"'{CONFIG_FILE}' optional 'source_dir' must be a non-empty string.")
    source_dir = source_dir.strip()
    if not os.path.isabs(source_dir):
        source_dir = os.path.join(script_dir, source_dir)

    source_file_prefixes = config.get("source_file_prefixes", {})
    if source_file_prefixes is None:
        source_file_prefixes = {}
    if not isinstance(source_file_prefixes, dict):
        abort(f"'{CONFIG_FILE}' optional 'source_file_prefixes' must be a JSON object.")

    normalized_prefixes = {}
    for key, prefix in source_file_prefixes.items():
        if key not in SOURCE_KEYS:
            abort(
                f"'{CONFIG_FILE}' source_file_prefixes contains unexpected key '{key}'."
            )
        if not isinstance(prefix, str) or not prefix.strip():
            abort(
                f"'{CONFIG_FILE}' source_file_prefixes entry for '{key}' must be a non-empty string."
            )
        normalized_prefixes[key] = prefix.strip()

    if not normalized_prefixes:
        normalized_prefixes = DEFAULT_SOURCE_FILE_PREFIXES.copy()

    missing_prefixes = set(SOURCE_KEYS) - set(normalized_prefixes.keys())
    if missing_prefixes:
        abort(
            f"'{CONFIG_FILE}' must define source_file_prefixes for: {sorted(missing_prefixes)}."
        )

    explicit_files = None
    files = config.get("files")
    if files is not None:
        if not isinstance(files, dict):
            abort(f"'{CONFIG_FILE}' optional 'files' must be a JSON object.")
        explicit_files = {}
        for key, filename in files.items():
            if not isinstance(filename, str) or not filename.strip():
                abort(
                    f"'{CONFIG_FILE}' entry for '{key}' must be a non-empty string filename."
                )
            explicit_files[key] = os.path.join(source_dir, filename.strip())

    column_map = config.get("column_map", {})
    if not isinstance(column_map, dict):
        abort(f"'{CONFIG_FILE}' optional 'column_map' must be a JSON object.")

    for key, mapping in column_map.items():
        if not isinstance(mapping, dict):
            abort(f"'{CONFIG_FILE}' column_map for '{key}' must be a JSON object.")
        for src_col, tpl_col in mapping.items():
            if not isinstance(src_col, str) or not isinstance(tpl_col, str):
                abort(
                    f"'{CONFIG_FILE}' column_map entry for '{key}' must map strings to strings."
                )

    organization_id = config.get("organization_id")
    if not isinstance(organization_id, str) or not organization_id.strip():
        abort(f"'{CONFIG_FILE}' must define 'organization_id' as a non-empty string.")
    organization_id = organization_id.strip()

    return source_dir, normalized_prefixes, explicit_files, column_map, organization_id


def load_batch_identifier_map(map_path):
    """Load group-to-batch mappings from a CSV file."""
    if not os.path.exists(map_path):
        abort(f"Batch map file not found: {os.path.abspath(map_path)}")

    df = pd.read_csv(map_path, dtype=str)
    if "group_id" not in df.columns or "batch_id" not in df.columns:
        abort(
            f"Batch map file '{os.path.basename(map_path)}' must contain columns 'group_id' and 'batch_id'."
        )

    mapping = {}
    for row_index, row in enumerate(df.itertuples(index=False), start=2):
        group_id = clean_str(row.group_id)
        batch_id = clean_str(row.batch_id)
        if not group_id or not batch_id:
            abort(
                f"Batch map file contains empty group_id or batch_id on row {row_index}."
            )
        if not batch_id.isdigit():
            abort(
                f"Batch ID '{batch_id}' for group '{group_id}' must be numeric."
            )
        if group_id in mapping:
            abort(f"Duplicate group_id '{group_id}' found in batch map file.")
        mapping[group_id] = batch_id
    return mapping


def parse_source_filename(filename, prefix):
    """Parse a source filename and return the group identifier if it matches."""
    if not filename.lower().endswith(".csv"):
        return None

    if not filename.startswith(prefix + "_"):
        return None

    stem = filename[:-4]
    remainder = stem[len(prefix) + 1 :]
    parts = remainder.split("_")
    if len(parts) < 3:
        return None

    return "_".join(parts[:2])


def discover_source_groups(source_dir, prefixes):
    """Discover available source groups in the source directory."""
    if not os.path.isdir(source_dir):
        abort(f"Source directory not found: {os.path.abspath(source_dir)}")

    groups = {}
    for filename in os.listdir(source_dir):
        for key, prefix in sorted(prefixes.items(), key=lambda item: -len(item[1])):
            group_id = parse_source_filename(filename, prefix)
            if not group_id:
                continue

            group_files = groups.setdefault(group_id, {})
            if key in group_files:
                abort(
                    f"Duplicate file detected for group '{group_id}' and key '{key}':"
                    f" '{filename}' and '{os.path.basename(group_files[key])}'."
                )
            group_files[key] = os.path.join(source_dir, filename)
            break

    return groups


def format_source_groups(groups):
    lines = []
    for group_id, files in sorted(groups.items()):
        status = "COMPLETE" if set(files.keys()) == set(SOURCE_KEYS) else "INCOMPLETE"
        lines.append(f"  {group_id}: {status} ({len(files)}/6 files)")
    return "\n".join(lines)


def normalize_column_name(name):
    """Normalize a column header for matching across source and template."""
    if name is None:
        return ""
    return "".join(ch.lower() for ch in str(name).strip() if ch.isalnum())


def find_template_header_row(ws):
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if not row:
            continue
        normalized = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if any("batch identifier" in cell for cell in normalized if cell):
            return [str(cell).strip() if cell is not None else "" for cell in row]

    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            return [str(cell).strip() if cell is not None else "" for cell in row]

    return []


def close_workbook_safe(wb):
    if wb is None:
        return

    archive = getattr(wb, "_archive", None)
    try:
        wb.close()
    except Exception:
        pass

    if archive is not None:
        try:
            if hasattr(archive, "fp") and archive.fp is not None and getattr(archive.fp, "closed", False):
                archive.fp = None
        except Exception:
            pass
        try:
            archive.close()
        except Exception:
            pass

    try:
        if hasattr(wb, "_archive"):
            wb._archive = None
    except Exception:
        pass


def load_template_headers(template_path):
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    headers = {}
    try:
        for key, sheet_name in SHEET_NAMES.items():
            if sheet_name not in wb.sheetnames:
                abort(
                    f"Template sheet '{sheet_name}' not found in workbook. "
                    f"Check SHEET_NAMES config."
                )
            ws = wb[sheet_name]
            header_row = find_template_header_row(ws)
            if not header_row:
                abort(f"Could not determine header row for template sheet '{sheet_name}'.")
            headers[key] = [cell for cell in header_row if cell]
            log(f"[{sheet_name}] Template header row loaded with {len(headers[key])} columns", "OK")
    finally:
        close_workbook_safe(wb)
    return headers


def get_source_key_for_target(target_key):
    return TARGET_SOURCE_KEY.get(target_key, target_key)


def build_column_map(frames, template_headers, source_paths=None, group_id=None):
    log("Building source-to-template column mappings ...", "HEADER")
    column_map = {}
    for key, sheet_name in SHEET_NAMES.items():
        source_key = get_source_key_for_target(key)
        source_path = source_paths.get(source_key) if source_paths else None
        context = format_context(group_id=group_id, file_path=source_path, key=key)
        df = frames.get(source_key)
        if df is None:
            log(
                f"[{key}] No source frame available for target sheet '{sheet_name}'. "
                f"Skipping mapping.", "WARN", context
            )
            continue

        template_cols = template_headers.get(key, [])
        template_lookup = {
            normalize_column_name(col): col
            for col in template_cols if col
        }
        mapping = {}
        unmapped = []
        for src_col in df.columns:
            norm_src = normalize_column_name(src_col)
            if norm_src in template_lookup:
                mapping[src_col] = template_lookup[norm_src]
            else:
                unmapped.append(src_col)
        if unmapped:
            log(
                f"[{key}] {len(unmapped)} source column(s) were not mapped to template "
                f"columns: {unmapped}", "WARN", context
            )

        template_unmapped = [col for col in template_cols if col and normalize_column_name(col) not in {
            normalize_column_name(fbdi_col) for fbdi_col in mapping.values()
        }]
        if template_unmapped:
            log(
                f"[{key}] {len(template_unmapped)} template column(s) were not mapped from source "
                f"columns: {template_unmapped}", "WARN", context
            )

        column_map[key] = mapping
        log(
            f"[{key}] mapped {len(mapping)}/{len(df.columns)} source columns "
            f"to template columns", "OK"
        )
    return column_map


def validate_column_map(column_map, frames, template_headers, source_paths=None, group_id=None):
    log("Validating configured column map ...", "HEADER")
    error_count = 0
    for key, mapping in column_map.items():
        source_key = get_source_key_for_target(key)
        source_df = frames.get(source_key)
        source_path = source_paths.get(source_key) if source_paths else None
        context = format_context(group_id=group_id, file_path=source_path, key=key)

        if source_df is None:
            log(
                f"Column map key exists in config but no source data is available for source '{source_key}'.",
                "ERROR",
                context,
            )
            error_count += 1
            continue

        template_cols = template_headers.get(key, [])
        if template_cols is None:
            log(
                f"Template headers not loaded for sheet '{SHEET_NAMES.get(key)}'.",
                "ERROR",
                context,
            )
            error_count += 1
            continue

        source_lookup = {
            normalize_column_name(col): col for col in source_df.columns if col
        }
        normalized_lookup = {
            normalize_column_name(col): col for col in template_cols if col
        }
        for src_col, tpl_col in mapping.items():
            if normalize_column_name(src_col) not in source_lookup:
                log(
                    f"Source column '{src_col}' from config column_map not found in CSV for source '{source_key}'.",
                    "ERROR",
                    context,
                )
                error_count += 1
            if normalize_column_name(tpl_col) not in normalized_lookup:
                log(
                    f"Template column '{tpl_col}' from config column_map not found in sheet '{SHEET_NAMES.get(key)}'.",
                    "ERROR",
                    context,
                )
                error_count += 1
    return error_count


def resolve_column_map_template_names(column_map, template_headers):
    """Resolve config/template mapping to exact header names in the workbook."""
    resolved = {}
    for key, mapping in column_map.items():
        template_cols = template_headers.get(key, [])
        normalized_lookup = {
            normalize_column_name(col): col for col in template_cols if col
        }
        resolved[key] = {
            src_col: normalized_lookup.get(normalize_column_name(tpl_col), tpl_col)
            for src_col, tpl_col in mapping.items()
        }
    return resolved


def validate_mandatory_star_fields(transformed, template_headers, group_id=None):
    log("Validating mandatory '*' template fields ...", "HEADER")
    error_count = 0
    for key, df in transformed.items():
        template_cols = template_headers.get(key, [])
        mandatory_cols = [col for col in template_cols if col.startswith("*")]
        if not mandatory_cols:
            continue

        context = format_context(group_id=group_id, key=key)
        output_lookup = {
            normalize_column_name(col): col for col in df.columns
        }

        for tpl_col in mandatory_cols:
            normalized = normalize_column_name(tpl_col)
            output_col = output_lookup.get(normalized)
            if output_col is None:
                log(
                    f"Mandatory template column '{tpl_col}' is missing from transformed output.",
                    "ERROR",
                    context,
                )
                error_count += 1
                continue

            blank_mask = df[output_col].apply(clean_str) == ""
            if blank_mask.any():
                excel_rows = [i + 2 for i in df.index[blank_mask].tolist()]
                log(
                    f"Mandatory column '{tpl_col}' has blank values in {len(excel_rows)} row(s): Excel rows {excel_rows}",
                    "ERROR",
                    context,
                )
                error_count += len(excel_rows)
            else:
                log(f"Mandatory column '{tpl_col}' all populated", "OK", context)
    return error_count


# ---------------------------------------------------------------------------
# FBDI SHEET NAMES — these must match exactly the sheet names in the Oracle
# template. Adjust here if your template version uses different names.
# ---------------------------------------------------------------------------
SHEET_NAMES = {
    "party"             : "HZ_IMP_PARTIES_T",
    "locations"         : "HZ_IMP_LOCATIONS_T",
    "party_sites"       : "HZ_IMP_PARTYSITES_T",
    "party_site_uses"   : "HZ_IMP_PARTYSITEUSES_T",
    "accounts"          : "HZ_IMP_ACCOUNTS_T",
    "account_sites"     : "HZ_IMP_ACCTSITES_T",
    "account_site_uses" : "HZ_IMP_ACCTSITEUSES_T",
    "profile_class"     : "RA_CUSTOMER_PROFILES_INT_ALL",
}

TARGET_SOURCE_KEY = {
    "party_site_uses"   : "party_sites",
    "account_site_uses" : "account_sites",
}

SOURCE_OSR_COLUMNS = {
    "party"        : "party_original_system_reference",
    "locations"    : "location_original_system_reference",
    "party_sites"  : "party_site_original_system_reference",
    "accounts"     : "customer_account_source_system_reference",
    "account_sites": "account_site_source_system_reference",
}

# ---------------------------------------------------------------------------
# COLUMN MAPPINGS — built dynamically from template headers and source CSV
# column names. The script matches source-to-template names using normalized
# values so underscores, punctuation and casing do not matter.
# ---------------------------------------------------------------------------
COLUMN_MAP = {}

# ---------------------------------------------------------------------------
# PARENT-CHILD REFERENCE RULES
# Defines which column in a child sheet must have matching values in the
# parent sheet's source key column.
# ---------------------------------------------------------------------------
REFERENCE_RULES = {
    "party_sites"  : [
        ("party_original_system_reference", "party"),
        ("location_original_system_reference", "locations"),
    ],
    "accounts"     : [
        ("party_original_system_reference", "party"),
    ],
    "account_sites": [
        ("customer_account_source_system_reference", "accounts"),
        ("party_site_orig_system_reference", "party_sites"),
    ],
    "profile_class": [
        ("customer_account_source_system_reference", "accounts"),
    ],
}


# =============================================================================
# HELPERS
# =============================================================================

def format_context(group_id=None, file_path=None, sheet_name=None, key=None):
    parts = []
    if group_id:
        parts.append(f"group={group_id}")
    if file_path:
        parts.append(f"file={os.path.basename(file_path)}")
    if sheet_name:
        parts.append(f"sheet={sheet_name}")
    elif key:
        parts.append(f"sheet_key={key}")
    return " ".join(parts) if parts else None


def log(msg, level="INFO", context=None):
    prefix = {
        "INFO"   : "  ",
        "OK"     : "  [OK]",
        "WARN"   : "  [WARN]",
        "ERROR"  : "  [ERROR]",
        "HEADER" : "",
    }.get(level, "  ")
    ts = datetime.now().strftime("%H:%M:%S")
    ctx = f" [{context}]" if context else ""
    print(f"[{ts}] {prefix}{ctx} {msg}")

def clean_str(val):
    """Return normalized string; empty string for NaN/None."""
    if pd.isna(val):
        return ""
    return " ".join(str(val).strip().split())


def is_date_column(header_name):
    """Return True for template columns whose name indicates a date field."""
    if not header_name:
        return False
    return bool(re.search(r"\bDate\b", str(header_name), flags=re.IGNORECASE))


def format_date_value(value):
    """Format populated date values as YYYY/MM/DD."""
    value = clean_str(value)
    if not value:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")

    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return value
    return parsed.strftime("%Y/%m/%d")


def abort(msg):
    log(msg, "ERROR")
    log("Script aborted. Fix the issues above and re-run.", "ERROR")
    sys.exit(1)


# =============================================================================
# STEP 1 — Parse and validate batch ID argument
# =============================================================================

def parse_arguments():
    parser = argparse.ArgumentParser(
        description=(
            "Prepare Oracle Fusion AR Customer FBDI from dynamic source files. "
            "When multiple complete source groups are found, a batch map can be used to "
            "automatically assign the correct batch identifier for each group."
        )
    )
    parser.add_argument(
        "--batch-id",
        help=(
            "Numeric FBDI batch identifier to use for all processed files when no "
            "batch map file is provided."
        ),
        dest="batch_id",
    )
    parser.add_argument(
        "--batch-map",
        help=(
            f"Path to group-to-batch map CSV file (default: {BATCH_MAP_FILE}). "
            "When present and multiple complete groups exist, the script will "
            "process each group using its mapped batch identifier."
        ),
        default=BATCH_MAP_FILE,
        dest="batch_map",
    )
    parser.add_argument(
        "--group",
        help=(
            "Source group to process, e.g. 0001_BPS. If omitted and a batch map is "
            "available, all complete groups are processed automatically."
        ),
        dest="group_id",
    )
    parser.add_argument(
        "--list-groups",
        action="store_true",
        help="List detected source groups and exit.",
    )
    parser.add_argument(
        "--parallel",
        action="store_true",
        help="Process all complete source groups in parallel",
    )
    parser.add_argument(
        "--max-workers",
        type=int,
        default=min(4, os.cpu_count() or 1),
        help="Maximum number of parallel workers when using --parallel",
    )
    return parser.parse_args()


def get_batch_id(batch_id):
    if not batch_id or not isinstance(batch_id, str):
        abort("Batch ID cannot be blank.")
    batch_id = batch_id.strip()
    if not batch_id:
        abort("Batch ID cannot be blank.")
    if not batch_id.isdigit():
        abort("Batch ID must be a numeric value for the Number field.")
    batch_id_num = int(batch_id)
    log(f"Batch ID: {batch_id_num}", "OK")
    return batch_id_num


def locate_template():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, TEMPLATE_DIR, TEMPLATE_FILENAME)
    if not os.path.exists(template_path):
        abort(
            f"FBDI template not found at: {os.path.abspath(template_path)}\n"
            f"        Please place '{TEMPLATE_FILENAME}' in the '{TEMPLATE_DIR}/' folder."
        )
    log(f"Template found: {template_path}", "OK")
    return template_path


def process_source_group(batch_id, source_files, template_path, config_column_map, organization_id, group_id=None):
    if group_id:
        log(f"Processing source group: {group_id}", "HEADER")
    else:
        log("Processing explicit source mapping", "HEADER")

    frames = read_sources(source_files)
    template_headers = load_template_headers(template_path)
    column_map = build_column_map(frames, template_headers)
    if config_column_map:
        for key, mapping in config_column_map.items():
            if key in column_map and isinstance(column_map[key], dict):
                column_map[key].update(mapping)
            else:
                column_map[key] = mapping
        log("Loaded column map from config file and merged with template-derived mappings.", "OK")

    if "party_site_uses" in column_map:
        column_map["party_site_uses"].pop("from_date", None)

    column_map = resolve_column_map_template_names(column_map, template_headers)
    error_count = validate_column_map(
        column_map,
        frames,
        template_headers,
        source_paths=source_files,
        group_id=group_id,
    )
    error_count += validate_orig_sys_refs(frames, source_paths=source_files, group_id=group_id)

    transformed = transform(
        frames,
        column_map,
        batch_id,
        template_headers,
        organization_id,
        source_paths=source_files,
        group_id=group_id,
    )
    error_count += validate_mandatory_star_fields(transformed, template_headers, group_id=group_id)

    output_path = write_to_template(template_path, transformed, batch_id, group_id)
    print_summary(batch_id, transformed, output_path, error_count, group_id)
    return output_path, error_count


# =============================================================================
# STEP 3 — Read all source CSV files
# =============================================================================

def read_sources(source_files):
    log("Reading source CSV files ...", "HEADER")
    frames = {}
    for key, path in source_files.items():
        if not os.path.exists(path):
            abort(f"Source file not found: {os.path.abspath(path)}")
        df = pd.read_csv(path, dtype=str)
        # Strip whitespace from all column headers and cell values
        df.columns = [c.strip() for c in df.columns]
        df = df.apply(lambda col: col.str.strip() if col.dtype == object else col)
        log(f"Loaded '{key}': {len(df)} rows | columns: {list(df.columns)}", "OK")
        frames[key] = df
    return frames


# =============================================================================
# STEP 4 — Validate OriginalSystemReference columns
# =============================================================================

def validate_orig_sys_refs(frames, source_paths=None, group_id=None):
    """Validate source reference key columns and parent-child relationships."""
    log("Validating source reference keys ...", "HEADER")
    error_count = 0

    # --- A) Not-blank check ---
    sheets_with_osr = ["party", "locations", "party_sites", "accounts", "account_sites"]

    for key in sheets_with_osr:
        df = frames[key]
        source_path = source_paths.get(key) if source_paths else None
        context = format_context(group_id=group_id, file_path=source_path, key=key)
        osr_col = SOURCE_OSR_COLUMNS.get(key)
        if not osr_col:
            continue

        if osr_col not in df.columns:
            log(
                f"Column '{osr_col}' not found. Check your CSV column names.",
                "ERROR",
                context,
            )
            error_count += 1
            continue

        blank_mask = df[osr_col].apply(clean_str) == ""
        blank_rows = df[blank_mask]
        if not blank_rows.empty:
            excel_rows = [i + 2 for i in blank_rows.index.tolist()]
            log(
                f"'{osr_col}' is blank in {len(blank_rows)} row(s): Excel rows {excel_rows}",
                "ERROR",
                context,
            )
            error_count += len(blank_rows)
        else:
            log(f"{osr_col} — all populated", "OK", context)

        osr_values = df[osr_col].apply(clean_str)
        dupes = osr_values[osr_values.duplicated(keep=False) & (osr_values != "")]
        if not dupes.empty:
            log(
                f"Duplicate {osr_col} values found: {dupes.unique().tolist()}",
                "ERROR",
                context,
            )
            error_count += len(dupes.unique())

    # --- B) Parent-child reference check ---
    log("Checking parent-child reference integrity ...", "HEADER")

    for child_key, rules in REFERENCE_RULES.items():
        child_df = frames[child_key]
        child_path = source_paths.get(child_key) if source_paths else None
        for fk_col, parent_key in rules:
            parent_df = frames[parent_key]
            parent_path = source_paths.get(parent_key) if source_paths else None
            parent_osr_col = SOURCE_OSR_COLUMNS.get(parent_key)
            context = format_context(group_id=group_id, file_path=child_path, key=child_key)

            if fk_col not in child_df.columns:
                log(
                    f"Foreign key column '{fk_col}' not found in CSV. Check your column names.",
                    "ERROR",
                    context,
                )
                error_count += 1
                continue

            if not parent_osr_col or parent_osr_col not in parent_df.columns:
                continue

            parent_osr_set = set(
                parent_df[parent_osr_col].apply(clean_str).tolist()
            ) - {""}

            child_fk_vals = child_df[fk_col].apply(clean_str)
            unmatched = child_df[
                (child_fk_vals != "") &
                (~child_fk_vals.isin(parent_osr_set))
            ]
            blank_fk = child_df[child_fk_vals == ""]

            if not unmatched.empty:
                excel_rows = [i + 2 for i in unmatched.index.tolist()]
                log(
                    f"'{fk_col}' has {len(unmatched)} value(s) that do not match any '{parent_key}' {parent_osr_col}: Excel rows {excel_rows}",
                    "ERROR",
                    context,
                )
                error_count += len(unmatched)
            else:
                log(
                    f"'{fk_col}' -> '{parent_key}' all references matched",
                    "OK",
                    context,
                )

            if not blank_fk.empty:
                excel_rows = [i + 2 for i in blank_fk.index.tolist()]
                log(
                    f"'{fk_col}' is blank in {len(blank_fk)} row(s): Excel rows {excel_rows}",
                    "ERROR",
                    context,
                )
                error_count += len(blank_fk)

    return error_count


# =============================================================================
# STEP 5 — Transform each source dataframe to FBDI column structure
# =============================================================================

def find_batch_identifier_column(template_cols):
    """Return the exact batch identifier header name from the template columns."""
    normalized_target = normalize_column_name("Batch Identifier")
    for col in template_cols:
        if normalize_column_name(col) == normalized_target:
            return col
    return "*Batch Identifier"


def find_organization_id_column(template_cols):
    """Return the exact organization ID header name from the template columns."""
    normalized_target = normalize_column_name("Organization ID")
    for col in template_cols:
        if normalize_column_name(col) == normalized_target:
            return col
    return None


def transform(frames, column_map, batch_id, template_headers, organization_id, source_paths=None, group_id=None):
    """
    Applies column mapping from the dynamically built mapping.
    Adds OPERATION_CODE = SYNC, the template's Batch Identifier field,
    and the configured Organization ID to every sheet.
    Source columns not in the mapping are dropped.
    """
    log("Transforming data to FBDI column structure ...", "HEADER")
    transformed = {}

    for key, sheet_name in SHEET_NAMES.items():
        source_key = get_source_key_for_target(key)
        source_path = source_paths.get(source_key) if source_paths else None
        context = format_context(group_id=group_id, file_path=source_path, key=key)
        df = frames.get(source_key)
        if df is None:
            log(
                f"[{key}] No source data found for target sheet '{sheet_name}'. Skipping.",
                "WARN",
                context,
            )
            continue

        mapping = column_map.get(key, {})
        out = pd.DataFrame(index=df.index)

        out["OPERATION_CODE"] = "SYNC"
        batch_col = find_batch_identifier_column(template_headers.get(key, []))
        out[batch_col] = batch_id

        org_col = find_organization_id_column(template_headers.get(key, []))
        if org_col:
            out[org_col] = organization_id

        for src_col, fbdi_col in mapping.items():
            if src_col in df.columns:
                out[fbdi_col] = df[src_col].apply(clean_str)
            else:
                log(
                    f"[{key}] Source column '{src_col}' not found in CSV — "
                    f"column '{fbdi_col}' will be blank in output.",
                    "WARN",
                    context,
                )
                out[fbdi_col] = ""

        if key == "account_sites":
            party_site_number_col = next(
                (col for col in out.columns if normalize_column_name(col) == normalize_column_name("Party Site Number")),
                None,
            )
            if party_site_number_col is not None:
                out[party_site_number_col] = ""

        for date_col in [col for col in out.columns if is_date_column(col)]:
            out[date_col] = out[date_col].apply(format_date_value)

        transformed[key] = out
        log(f"[{key}] Transformed: {len(out)} rows, {len(out.columns)} columns", "OK")

    return transformed


# =============================================================================
# STEP 6 — Write transformed data into the FBDI template
# =============================================================================

def write_to_template(template_path, transformed, batch_id, group_id=None):
    """
    Copies the Oracle FBDI .xlsm template to the output folder,
    then writes each transformed DataFrame into the corresponding sheet.

    Strategy: finds the header row in each sheet (the row containing
    'OriginalSystemReference' or the first column header), then writes
    data starting from the row immediately below it.
    Existing data rows in the template (if any) are cleared first.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Build output filename: CustomerImportTemplate_BATCH001[_0001_BPS].xlsm
    base_name = os.path.splitext(TEMPLATE_FILENAME)[0]
    output_name = f"{base_name}_{batch_id}"
    if group_id:
        output_name = f"{output_name}_{group_id}"
    output_name += ".xlsm"
    output_path = os.path.join(OUTPUT_DIR, output_name)

    # Copy template to output — never modify the original
    shutil.copy2(template_path, output_path)
    log(f"Template copied to: {output_path}", "OK")

    # Open the copied workbook preserving macros
    wb = openpyxl.load_workbook(output_path, keep_vba=True)

    try:
        log("Writing data into FBDI template sheets ...", "HEADER")

        for key, fbdi_df in transformed.items():
            sheet_name = SHEET_NAMES.get(key)
            if not sheet_name:
                continue

            if sheet_name not in wb.sheetnames:
                context = format_context(group_id=group_id, file_path=output_path, sheet_name=sheet_name)
                log(
                    f"Sheet '{sheet_name}' not found in template. "
                    f"Check SHEET_NAMES config. Skipping.", "WARN",
                    context,
                )
                continue

            ws = wb[sheet_name]

            # --- Find the header row ---
            # Oracle FBDI templates typically have instruction rows at the top
            # before the actual column header row. We scan for the header row by
            # matching normalized header names from the DataFrame against the row.
            header_row_num = None
            target_names = {
                normalize_column_name("OriginalSystemReference"),
                normalize_column_name("OPERATION_CODE"),
            }
            if len(fbdi_df.columns) > 2:
                target_names.add(normalize_column_name(fbdi_df.columns[2]))
            target_names.update(normalize_column_name(col) for col in fbdi_df.columns)

            for row in ws.iter_rows():
                for cell in row:
                    if not cell.value:
                        continue
                    if normalize_column_name(cell.value) in target_names:
                        header_row_num = cell.row
                        break
                if header_row_num:
                    break

            if not header_row_num:
                # Fallback: assume row 1 is header (simple template)
                header_row_num = 1
                context = format_context(group_id=group_id, file_path=output_path, sheet_name=sheet_name)
                log(
                    f"[{sheet_name}] Could not detect header row — "
                    f"defaulting to row 1.", "WARN",
                    context,
                )

            # --- Read template headers to build column position map ---
            # Keep the first occurrence of duplicate header names so repeated
            # fields like 'Address Line 1' map to the leftmost template column.
            template_headers = {}
            for cell in ws[header_row_num]:
                if not cell.value:
                    continue
                header_name = str(cell.value).strip()
                if header_name and header_name not in template_headers:
                    template_headers[header_name] = cell.column

            # --- Clear any existing data rows below the header ---
            data_start_row = header_row_num + 1
            for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

            # --- Write DataFrame rows into the template ---
            rows_written = 0
            for df_row_idx, df_row in fbdi_df.iterrows():
                excel_row = data_start_row + rows_written
                for fbdi_col, value in df_row.items():
                    col_num = template_headers.get(fbdi_col)
                    if col_num:
                        ws.cell(row=excel_row, column=col_num, value=value if value != "" else None)
                rows_written += 1

            log(f"[{sheet_name}] {rows_written} rows written", "OK")

        wb.save(output_path)
        log(f"Workbook saved: {output_path}", "OK")
        return output_path
    finally:
        close_workbook_safe(wb)


# =============================================================================
# STEP 7 — Print final summary
# =============================================================================

def print_summary(batch_id, transformed, output_path, error_count, group_id=None):
    log("", "HEADER")
    log("=" * 62, "HEADER")
    title = f" FBDI PREPARATION SUMMARY — Batch: {batch_id}"
    if group_id:
        title += f" | Group: {group_id}"
    log(title, "HEADER")
    log("=" * 62, "HEADER")

    total = 0
    for key, df in transformed.items():
        sheet = SHEET_NAMES.get(key, key)
        log(f"  {sheet:<35} {len(df):>5} rows", "HEADER")
        total += len(df)

    log(f"  {'TOTAL':<35} {total:>5} rows", "HEADER")
    log("=" * 62, "HEADER")

    if error_count > 0:
        log(
            f"  {error_count} validation issue(s) were found and logged above.",
            "HEADER"
        )
        log(
            "  The output file was still produced — review warnings before uploading.",
            "HEADER"
        )
    else:
        log("  All validations passed — no issues found.", "HEADER")

    log("", "HEADER")
    log(f"  Output file: {os.path.abspath(output_path)}", "HEADER")
    log("", "HEADER")
    log("  Next steps:", "HEADER")
    log("  1. Open the output .xlsm and spot-check a few rows", "HEADER")
    log("  2. In Fusion: Tools -> File Import and Export -> Upload the .xlsm", "HEADER")
    log("  3. Run scheduled process: Import Trading Community Data", "HEADER")
    log("  4. Review ESS log and error report in Fusion", "HEADER")
    log("=" * 62, "HEADER")


# =============================================================================
# MAIN
# =============================================================================

def main():
    args = parse_arguments()
    batch_id_arg = args.batch_id
    batch_map_path = args.batch_map
    group_arg = args.group_id

    batch_map = None
    if os.path.exists(batch_map_path):
        batch_map = load_batch_identifier_map(batch_map_path)
    elif batch_id_arg is None:
        abort(
            f"Batch map file not found and no --batch-id provided. Expected batch map at '{batch_map_path}'."
        )

    log("=" * 62, "HEADER")
    log(" Oracle Fusion AR — Customer FBDI Preparation  v2", "HEADER")
    log("=" * 62, "HEADER")

    # Step 2: Locate template
    template_path = locate_template()

    # Step 3: Load source file configuration and discover source groups
    source_dir, source_prefixes, explicit_files, config_column_map, organization_id = load_source_file_config()

    if explicit_files is not None:
        if group_arg:
            abort("The --group option cannot be used when explicit files are defined in the config.")
        if args.parallel:
            abort("Parallel processing is not supported when explicit files are provided in the config.")
        if args.list_groups:
            abort("Source group discovery is disabled when explicit files are provided in the config.")
        if batch_id_arg is None:
            abort("Explicit source files require --batch-id when no batch map file is used.")

        batch_id = get_batch_id(batch_id_arg)
        process_source_group(batch_id, explicit_files, template_path, config_column_map, organization_id)
        return

    groups = discover_source_groups(source_dir, source_prefixes)
    if not groups:
        abort(f"No source files found in '{os.path.abspath(source_dir)}'.")

    if args.list_groups:
        log("Detected source groups:", "HEADER")
        if batch_map is not None:
            print("Group ID | Batch ID | Status")
            for group_id, files in sorted(groups.items()):
                status = "COMPLETE" if set(files.keys()) == set(SOURCE_KEYS) else "INCOMPLETE"
                batch_id = batch_map.get(group_id, "<missing>")
                print(f"  {group_id:<10} | {batch_id:<10} | {status}")
        else:
            print(format_source_groups(groups))
        return

    complete_groups = {
        group_id: files
        for group_id, files in groups.items()
        if set(files.keys()) == set(SOURCE_KEYS)
    }

    if args.parallel:
        if group_arg:
            abort("The --group option cannot be used together with --parallel.")
        if not complete_groups:
            abort("No complete source groups available for parallel processing.")
        if batch_map is None:
            abort("Parallel processing requires a batch map file with one batch identifier per group.")

        log(
            f"Processing {len(complete_groups)} complete source groups in parallel with "
            f"{args.max_workers} workers.",
            "HEADER"
        )
        missing = [g for g in complete_groups if g not in batch_map]
        if missing:
            abort(
                f"Batch identifiers missing for groups: {missing}. "
                f"Update '{batch_map_path}' with these group IDs."
            )

        with concurrent.futures.ProcessPoolExecutor(max_workers=args.max_workers) as executor:
            futures = {
                executor.submit(
                    process_source_group,
                    get_batch_id(batch_map[group_id]),
                    files,
                    template_path,
                    config_column_map,
                    organization_id,
                    group_id,
                ): group_id
                for group_id, files in complete_groups.items()
            }
            for future in concurrent.futures.as_completed(futures):
                group_id = futures[future]
                try:
                    future.result()
                except Exception as exc:
                    abort(f"Processing group '{group_id}' failed: {exc}")
        return

    if group_arg:
        if group_arg not in complete_groups:
            abort(
                f"Requested group '{group_arg}' not found or is incomplete."
                f" Available complete groups:\n{format_source_groups(complete_groups)}"
            )
        source_files = complete_groups[group_arg]
        group_id = group_arg
    elif len(complete_groups) == 1:
        group_id, source_files = next(iter(complete_groups.items()))
        log(f"Automatically selected source group: {group_id}", "OK")
    elif batch_map is not None:
        missing = [g for g in complete_groups if g not in batch_map]
        if missing:
            abort(
                f"Batch identifiers missing for groups: {missing}. "
                f"Update '{batch_map_path}' with these group IDs."
            )

        total_errors = 0
        for group_id, source_files in sorted(complete_groups.items()):
            batch_id = get_batch_id(batch_map[group_id])
            _, errors = process_source_group(
                batch_id,
                source_files,
                template_path,
                config_column_map,
                organization_id,
                group_id,
            )
            total_errors += errors

        if total_errors:
            abort(f"Processing completed with {total_errors} validation issue(s). See logs above.")
        return
    else:
        abort(
            "Multiple complete source groups were found. Specify one with --group or provide a batch map file."
            f"\nAvailable groups:\n{format_source_groups(complete_groups)}"
        )

    if batch_map is not None:
        if group_id not in batch_map:
            abort(
                f"Batch identifier for group '{group_id}' was not found in '{batch_map_path}'."
            )
        batch_id = get_batch_id(batch_map[group_id])
    elif batch_id_arg is not None:
        batch_id = get_batch_id(batch_id_arg)
    else:
        abort(
            "No batch identifier available. Provide --batch-id or a batch map file."
        )

    frames = read_sources(source_files)

    # Step 4: Build mappings and validate source references
    template_headers = load_template_headers(template_path)
    column_map = build_column_map(frames, template_headers)
    if config_column_map:
        for key, mapping in config_column_map.items():
            if key in column_map and isinstance(column_map[key], dict):
                column_map[key].update(mapping)
            else:
                column_map[key] = mapping
        log("Loaded column map from config file and merged with template-derived mappings.", "OK")

    if "party_site_uses" in column_map:
        column_map["party_site_uses"].pop("from_date", None)

    column_map = resolve_column_map_template_names(column_map, template_headers)

    error_count = validate_column_map(
        column_map,
        frames,
        template_headers,
        source_paths=source_files,
        group_id=group_id,
    )
    error_count += validate_orig_sys_refs(frames, source_paths=source_files, group_id=group_id)

    # Step 5: Transform to FBDI column structure
    transformed = transform(frames, column_map, batch_id, template_headers, organization_id)
    error_count += validate_mandatory_star_fields(transformed, template_headers, group_id=group_id)

    # Step 6: Write into template copy
    output_path = write_to_template(template_path, transformed, batch_id, group_id)

    # Step 7: Summary
    print_summary(batch_id, transformed, output_path, error_count, group_id)


if __name__ == "__main__":
    main()